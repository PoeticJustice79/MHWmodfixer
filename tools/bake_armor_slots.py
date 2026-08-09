"""One-time baker: converts the community armor-slot physics reference
spreadsheet ("MHWs 방어구ID" -- community-compiled, originally by
Quaysar/RayVVV, Korean translation by 몬붕이) into the compact
tools/armor_slots_ch03.json.gz bundled with the app, powering the
"적용 방어구 변경" (retarget) tab's compatibility list.

Only the objective columns are baked (slot number, armor name, variant,
per-piece physics tokens, slinger presence) -- the spreadsheet's personal
mod-pack annotation columns are deliberately dropped.

Per-language display names are sourced from the CURRENTLY INSTALLED
GAME's own localization data (natives/stm/gamedesign/text/excel_equip/
armorseries.msg.23 -- see tools/msg_reader.py), not translated by hand --
that file already carries every armor series' real name in every
language the game itself ships, keyed by the exact same names this
sheet uses. This makes the previous session's hand-built, partial
NAME_EN_OVERRIDES table (English-only, ~86/104 entries, several
corrected-after-the-fact guesses) obsolete for anything the game data
actually covers.

Usage:  python tools/bake_armor_slots.py "<path to the .xlsx>" ["<game dir>"]
        (game dir defaults to auto_fix.DEFAULT_GAME_DIR; pass "" to skip
        the game-data pass entirely and bake with Korean names only)
"""
from __future__ import annotations

import gzip
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE.parent))  # for game_archive/auto_fix imports

OUT_PATH = _HERE / "armor_slots_ch03.json.gz"

# Column layout in the source sheet (1-based): B=set number, C=name,
# D=variant, E..I=pieces 1-5 (arm/body/helm/leg/waist), J=slinger √/×.
COL_SET, COL_NAME, COL_VARIANT = 2, 3, 4
COL_PIECES = [5, 6, 7, 8, 9]
COL_SLINGER = 10
DATA_START_ROW = 5

KNOWN_TOKENS = {"clsp", "chain", "gpuc"}

# i18n.py language code -> via.Language numeric code used inside the
# game's own .msg files (see tools/msg_reader.LANGUAGE_NAMES).
_LANG_CODE_MAP = {"en": 1, "ja": 0, "zh_tw": 12, "zh_cn": 13}

# This sheet's own name text sometimes differs cosmetically from the
# game's own ArmorSeries.msg spelling for the exact same set (a missing
# space, a compound word the sheet appends that the game's own name
# doesn't carry, a single-syllable spelling variant) -- confirmed by
# direct fuzzy lookup against the real game data, not guessed. Mapped
# here so those entries still get REAL game-sourced translations (all
# 4 languages) instead of falling back to Korean or an English-only
# manual override.
KO_NAME_ALIASES = {
    "수호룡세크레트": "수호룡 세크레트",  # sheet drops the space
    "실드후드": "실드",                    # sheet appends "후드"(hood); game's own name is bare
    "블로썸": "블로섬",                    # single-syllable spelling variant
    "길드 크로스": "길드크로스",            # sheet adds a space the game's name doesn't have
}

# Manually-confirmed English names for entries the game's ArmorSeries.msg
# does NOT cover -- unique/DLC accessory items (glasses, earrings, wigs)
# that aren't "armor series" at all. Sourced either directly from the
# user (playing the actual game) or a real name recognized on sight, NOT
# guessed -- kept minimal and English-only (no ja/zh source available for
# these, and per the user, en-only is an acceptable fallback for
# languages this project has no way to verify).
ACCESSORY_NAME_OVERRIDES = {
    "고우키": "Akuma",  # confirmed by the user directly -- MH's Street Fighter collab set
    "병사의 갑주(디럭스)": "Feudal Soldier",  # confirmed by the user directly (Deluxe Edition bonus set)
    "길드나이트(사전예약)": "Guild Knight",
    "깃 한 가닥 목걸이": "Pinion Necklace",
    "모험의 호크하트": "Hawkheart",
    "봉인의 안대": "Sealed Eyepatch", "봉인의 용해포": "Sealed Dragon Cloth",
    "스퀘어글라스": "Square Glasses", "언더림글라스": "Half Rim Glasses",
    "옷1": "Innerwear", "옷2": "Innerwear", "옷3": "Innerwear", "옷4": "Innerwear",
    "용왕의 척안": "Dragonking's Third Eye",
    "지략의 안경": "Strategist Spectacles",
    "하트글라스": "Lovely Shades",
    "헌신의 피어스": "Earrings of Dedication", "대식가의 귀걸이": "Gourmand's Earring",
    "노블레스": "Noblesse",
    "스자의 허리띠": "Suja's Belt",
}


_TIER_SUFFIXES = ("α", "β", "γ")


def _load_armor_series_names(game_dir: str) -> dict[str, dict[str, str]]:
    """{korean_name: {"en": ..., "ja": ..., "zh_tw": ..., "zh_cn": ...}},
    read directly from the currently installed game's own ArmorSeries.msg
    localization data. Returns {} (not an error) if no game dir is given
    or the file can't be read -- callers fall back to Korean-only names
    rather than fail the whole bake.

    The game's own data names armor by SPECIFIC RANK TIER (a monster with
    no low-rank set at all -- confirmed for several elder-dragon-tier
    monsters, e.g. Arkveld, Gore Magala, Rathalos -- has ONLY "이름α"/
    "이름β"/etc entries, never a bare "이름"). This project's own sheet
    names a whole SERIES with one flat name regardless of tier, so a bare
    Korean name with no exact msg-file match falls back to its α-tier
    entry (Korean suffix appended directly, "이름α"; the corresponding
    English/ja/zh values carry the suffix space-separated, "Name α" --
    confirmed by direct inspection, not assumed) with the tier marker
    stripped back off before use."""
    if not game_dir:
        return {}
    try:
        from game_archive import GameArchive
        from msg_reader import parse_msg
    except ImportError:
        return {}
    try:
        game = GameArchive(game_dir, log=lambda s: None)
        data = game.read_path("natives/stm/gamedesign/text/excel_equip/armorseries.msg.23")
        if data is None:
            return {}
        result = parse_msg(data)
    except Exception as exc:
        print(f"  [warn] could not read ArmorSeries.msg from game data: {exc}")
        return {}

    langs = result["languages"]
    try:
        ko_idx = langs.index(11)
    except ValueError:
        return {}
    lang_idx = {code: langs.index(via_code) for code, via_code in _LANG_CODE_MAP.items() if via_code in langs}

    bare: dict[str, dict[str, str]] = {}
    alpha_tier: dict[str, dict[str, str]] = {}  # base name (suffix stripped) -> names (suffix stripped)
    for e in result["entries"]:
        content = e["content"]
        ko = content[ko_idx].strip() if ko_idx < len(content) else ""
        if not ko or ko == "-":
            continue
        names = {code: content[i].strip() for code, i in lang_idx.items() if i < len(content) and content[i].strip()}
        if not names:
            continue
        if ko.endswith(_TIER_SUFFIXES):
            if not ko.endswith("α"):
                continue  # only α (the lowest/first tier) stands in as the series' representative name
            base_ko = ko[:-1]
            if base_ko not in alpha_tier:
                stripped = {code: v[:-1].strip() if v.endswith("α") else v for code, v in names.items()}
                alpha_tier[base_ko] = stripped
        else:
            bare[ko] = names

    for base_ko, names in alpha_tier.items():
        bare.setdefault(base_ko, names)
    return bare


def _parse_piece_cell(value) -> list[str] | None:
    """Returns the list of physics tokens for one piece, or None when the
    cell carries no usable per-piece data. Cells like "chain X"/"chain O"
    (shorthand rows the sheet uses for accessories/simple sets, not tied
    to a specific piece) also yield None -- those slots get an "unknown"
    profile and are excluded from automatic compatibility matching rather
    than guessed at."""
    if value is None:
        return None
    tokens = str(value).replace(" ", " ").split()
    if any(t in ("X", "O", "x", "o") for t in tokens):
        return None
    out = [t for t in tokens if t.lower() in KNOWN_TOKENS]
    return out if out else None


def _assign_genders(slots: dict[str, dict]) -> None:
    """Fills in each entry's "gender" field ('male'/'female'/None), in
    place. The sheet's 번호2 (variant) column doesn't carry an explicit
    gender marker -- per the user, a variant ending in 0 is the 남성
    (male) cut and the SIBLING variant ending in 1 (same set, same
    leading digits) is the 여성 (female) cut of the identical armor.
    Only labeled when BOTH siblings actually exist for that set --
    single-variant entries (accessories like glasses/necklaces, e.g.
    "089 깃 한 가닥 목걸이" variant 000 alone) are NOT a male/female pair
    and are deliberately left unlabeled rather than guessed at from the
    trailing digit alone, which would mislabel every ungendered
    accessory as "male"."""
    by_set: dict[str, set[str]] = {}
    for v in slots.values():
        by_set.setdefault(v["set"], set()).add(v["variant"])
    for v in slots.values():
        variant, set_no = v["variant"], v["set"]
        if not variant or variant[-1] not in ("0", "1"):
            v["gender"] = None
            continue
        sibling = variant[:-1] + ("1" if variant[-1] == "0" else "0")
        if sibling in by_set.get(set_no, ()):
            v["gender"] = "male" if variant[-1] == "0" else "female"
        else:
            v["gender"] = None


def bake(xlsx_path: Path, game_dir: str = "") -> tuple[dict, int]:
    """Returns (slots, game_names_matched_count)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    game_names = _load_armor_series_names(game_dir)
    matched = 0

    slots: dict[str, dict] = {}
    current_set = None
    current_name = None
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        cells = {c.column: c.value for c in row}
        set_no = cells.get(COL_SET)
        name = cells.get(COL_NAME)
        variant = cells.get(COL_VARIANT)
        if set_no is not None:
            current_set = str(set_no).strip()
        if name is not None:
            current_name = str(name).strip()
        if current_set is None or variant is None:
            continue
        variant = str(variant).strip()
        if not variant.isdigit():
            continue

        pieces = {}
        any_data = False
        for i, col in enumerate(COL_PIECES, start=1):
            tokens = _parse_piece_cell(cells.get(col))
            if tokens is not None:
                pieces[str(i)] = tokens
                any_data = True
        slinger_raw = cells.get(COL_SLINGER)
        slinger = None
        if slinger_raw is not None:
            s = str(slinger_raw).strip()
            slinger = True if "√" in s else False if s in ("×", "x", "X") else None

        ko_name = current_name or "?"
        names = game_names.get(ko_name) or game_names.get(KO_NAME_ALIASES.get(ko_name, ""))
        if names:
            matched += 1
        elif ko_name in ACCESSORY_NAME_OVERRIDES:
            names = {"en": ACCESSORY_NAME_OVERRIDES[ko_name]}
        else:
            names = {}

        slots[f"{current_set}/{variant}"] = {
            "set": current_set,
            "variant": variant,
            "name": ko_name,
            "names": names,  # {"en"/"ja"/"zh_tw"/"zh_cn": str, ...} -- missing key => fall back to Korean
            "pieces": pieces if any_data else None,  # None => profile unknown
            "slinger": slinger,
        }
    _assign_genders(slots)
    return slots, matched


def main():
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        sys.exit(1)
    xlsx_path = Path(sys.argv[1])
    if len(sys.argv) == 3:
        game_dir = sys.argv[2]
    else:
        from auto_fix import DEFAULT_GAME_DIR
        game_dir = DEFAULT_GAME_DIR if Path(DEFAULT_GAME_DIR).is_dir() else ""

    slots, matched = bake(xlsx_path, game_dir)
    known = sum(1 for v in slots.values() if v["pieces"])
    with_names = sum(1 for v in slots.values() if v["names"])
    payload = {"_meta": {"source": "community armor-slot physics reference (personal columns stripped) "
                                    "+ game's own ArmorSeries.msg localization data",
                          "entries": len(slots), "with_profile": known, "with_names": with_names},
               "slots": slots}
    with gzip.open(OUT_PATH, "wt", encoding="utf-8", compresslevel=9) as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"baked {len(slots)} slot-variants ({known} with a usable physics profile, "
          f"{with_names} with a translated name [{matched} from live game data]) -> {OUT_PATH}")


if __name__ == "__main__":
    main()
